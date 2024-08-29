import serial
import time
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
import subprocess
import customtkinter as ctk
import threading
import re
import sys
import glob
import queue

class APNamerGUI(ctk.CTk):
    WIDTH = 800
    HEIGHT = 900 

    def __init__(self):
        super().__init__()
        self.title("AP Namer")
        self.initialize_main_window()
        self.queue = queue.Queue()
        self.after(100, self.process_queue)

    def initialize_main_window(self):
        self.geometry(f"{APNamerGUI.WIDTH}x{APNamerGUI.HEIGHT}")
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        self.tabControl = ctk.CTkTabview(self)
        self.tabControl.pack(expand=1, fill="both")

        self.tab_1 = self.tabControl.add("Info")
        self.tab_2 = self.tabControl.add("Provision")
        self.tab_3 = self.tabControl.add("Manual Port")

        self.com_port = None
        self.file_path = None
        self.disabled = False
        self.override_used = False

        self.create_info_tab()
        self.create_provision_tab()
        self.create_manual_port_tab()

        self.com_port_thread = threading.Thread(target=self.check_com_ports)
        self.com_port_thread.daemon = True
        self.com_port_thread.start()

    def create_info_tab(self):
        intro_text = """
        Welcome to AP Namer!

        -Click to provisioning tab

        1. Plug in serial cable to computer
        2. Select Excel file in next tab
        3. Plug serial cable into AP
        4. Follow prompts, plug AP into power
        5. Click restart when you want to provision the next AP, it will use the SAME port
        
        *Use the manual port tab if AruProTool is not finding the COM port you have plugged in
        """

        intro_textbox = ctk.CTkTextbox(self.tab_1, wrap=tk.WORD, height=20)
        intro_textbox.insert(tk.END, intro_text)
        intro_textbox.configure(state=tk.DISABLED)
        intro_textbox.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

    def create_provision_tab(self):
        self.excel_button = ctk.CTkButton(self.tab_2, text="Select Excel File", command=self.select_excel_file)
        self.excel_button.pack(pady=4)
        
        self.restart_button = ctk.CTkButton(self.tab_2, text="Next AP", command=self.restart)
        self.restart_button.pack(pady=4)
        
        self.output_textbox = ctk.CTkTextbox(self.tab_2, wrap=tk.WORD, height=20)
        self.output_textbox.configure(state=tk.NORMAL)
        self.output_textbox.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

    def create_manual_port_tab(self):
        manual_override_text = """
        This is for if the program is trying to use the incorrect COM port or one that is already in use.

        Windows users should type COMx, eg. COM1,COM2 - all uppercase, no whitespaces.

        This does not forcefully open a COM port, rather it allows you to select a port manually. Check device manager to see 
        appropriate ports named:
        USB Serial...(COMx) --AP 5xx series 
        or
        Prolific USB-to-Serial Comm Port(COMx) --AP 3xx series
        """

        manual_override_textbox = ctk.CTkTextbox(self.tab_3, wrap=tk.WORD, height=10)
        manual_override_textbox.insert(tk.END, manual_override_text)
        manual_override_textbox.configure(state=tk.DISABLED)
        manual_override_textbox.pack(fill=tk.BOTH, expand=True, padx=20, pady=(20, 0), side=tk.TOP)

        override_frame = ctk.CTkFrame(self.tab_3)
        override_frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

        self.override_button = ctk.CTkButton(override_frame, text="Override", command=self.input)
        self.override_button.pack()

    def process_queue(self):
        try:
            message = self.queue.get(0)
            self.update_output(message)
        except queue.Empty:
            pass
        self.after(100, self.process_queue)

    def toggle_buttons(self):
        self.disabled = not self.disabled
        state = tk.DISABLED if self.disabled else tk.NORMAL
        self.restart_button.configure(state=state)
        self.excel_button.configure(state=state)
        self.override_button.configure(state=state)
    
        # Disable/enable widgets within tabs instead of the tabs themselves
        for widget in self.tab_1.winfo_children() + self.tab_2.winfo_children() + self.tab_3.winfo_children():
            if isinstance(widget, (ctk.CTkButton, ctk.CTkEntry, ctk.CTkTextbox)):
                widget.configure(state=state)

        # Special handling for the output textbox
        if state == tk.DISABLED:
            self.output_textbox.configure(state=tk.NORMAL)
        else:
            self.output_textbox.configure(state=tk.DISABLED)
    def input(self):
        dialog = ctk.CTkInputDialog(text="Type in COM port", title="Override Dialog")
        self.com_port = dialog.get_input()
        self.override_used = True
        self.queue.put(f"Overridden, you are now using: {self.com_port}")

    def restart(self):
        self.output_textbox.delete(1.0, tk.END)
        if self.com_port is None:
            self.input()
        if self.file_path is None:
            self.select_excel_file()
        if self.com_port and self.file_path:
            threading.Thread(target=self.start_serial, args=(self.com_port, self.file_path)).start()

    def select_excel_file(self):
        file_path = filedialog.askopenfilename(title="Select Excel file", filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            self.output_textbox.delete(1.0, tk.END)
            self.queue.put(f"Selected Excel file: {file_path}")
            self.file_path = file_path
            self.process_excel_file(file_path)

    def update_output(self, message):
        self.output_textbox.insert(tk.END, message + '\n')
        self.output_textbox.see(tk.END)

    def check_com_ports(self):
        if self.override_used:
            return
        first_instance_find_com = True
        first_instance_no_com = True
        while True:
            self.com_port = self.find_com_port()
            if self.com_port and first_instance_find_com:
                self.queue.put(f"COM Port found: {self.com_port}")
                first_instance_find_com = False
                break
            if first_instance_no_com:
                self.queue.put("No COM Port found")
                first_instance_no_com = False
            time.sleep(1)

    def process_excel_file(self, file_path):
        if self.com_port:
            threading.Thread(target=self.start_serial, args=(self.com_port, file_path)).start()

    def find_com_port(self):
        if sys.platform.startswith('win32'):
            powershell_cmd = [
            'powershell.exe',
            '-Command',
            '''
            $5xxcomPort = Get-CimInstance Win32_PnPEntity | Where-Object { $_.Caption -like '*USB Serial Port*' }
            $3xxcomPort = Get-CimInstance Win32_PnPEntity | Where-Object { $_.Caption -like '*Prolific USB-to-Serial Comm Port*' }

            if ($5xxcomPort) {
                $5xxcomPortName = $5xxcomPort.Name
                Write-Output "USB Serial Port found on COM port: $5xxcomPortName"
            } elseif ($3xxcomPort) {
                $3xxcomPortName = $3xxcomPort.Name
                Write-Output "USB Serial Port found on COM port: $3xxcomPortName"
            } else {
                Write-Output "USB Serial Port not found"
            }
            '''
            ]
            startupinfo = subprocess.STARTUPINFO()
            startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
            try:
                result = subprocess.check_output(powershell_cmd, text=True, startupinfo=startupinfo)
                lines = result.strip().split('\n')
                for line in lines:
                    if "USB Serial Port" in line:
                        com_port_index = line.find("(COM")
                        if com_port_index != -1:
                            com_port = line[com_port_index:].strip("() ")
                            return com_port
                else:
                    return None
            except subprocess.CalledProcessError as e:
                print(f"Error: {e}")
                return None
        else:
            port_list = glob.glob('/dev/ttyUSB*')
            if not port_list:
                self.queue.put("No COM port found")
            else:
                return port_list[0]

    def process_mac(self, mac_to_match, com_port, wb):
        sheet = wb.active
        headers = {}
        for cell in sheet[1]:
            if cell.value:
                headers[cell.value] = cell.column

        mac_column_index = headers.get("MAC")
        name_column_index = headers.get("AP Name")
        group_column_index = headers.get("AP Group")

        if mac_column_index is None or name_column_index is None or group_column_index is None:
            self.queue.put("Error: Required column not found in Excel file.")
            return

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[mac_column_index - 1] == mac_to_match and row[mac_column_index - 1] is not None:
                name = row[name_column_index - 1]
                group = row[group_column_index - 1]

                with serial.Serial(com_port, baudrate=9600, timeout=1, parity=serial.PARITY_NONE, bytesize=8, stopbits=1, xonxoff=False, rtscts=False, dsrdtr=False) as ser:
                    self.queue.put(f"You are using COM port: {com_port}")

                    commands = [
                        "purgeenv\r\n",
                        "saveenv\r\n",
                        f"set name {name}\r\n",
                        f"set group {group}\r\n",
                        "saveenv \r\n",
                        "printenv \r\n"
                    ]

                    for cmd in commands:
                        ser.write(cmd.encode())
                        time.sleep(1)

                    self.queue.put("Sent commands...")
                    time.sleep(2)

                    response = ser.read_all().decode()

                    actualName = re.search(r'\bname=([^\n]+)', response)
                    actualGroup = re.search(r'\bgroup=([^\n]+)', response)

                    self.queue.put(f"Extracted name: {actualName.group(1) if actualName else 'Not found'} Extracted group: {actualGroup.group(1) if actualGroup else 'Not found'}")
                    time.sleep(1)
                    self.queue.put("\n\nAP Successfully Provisioned.")
                break
        else:
            self.queue.put("MAC address not found in the Excel file.")

    def start_serial(self, com_port, file_path):
        self.toggle_buttons()
        try:
            wb = load_workbook(file_path, read_only=True)
            self.file_path = file_path

            with serial.Serial(com_port, baudrate=9600, timeout=1) as ser:
                no_data_prompt_printed = False

                while True:
                    data = ser.readline()

                    if b"Hit <Enter>" in data:
                        self.queue.put(data.decode())
                        ser.write(b"\r\n")
                        break
                    elif not data:
                        if not no_data_prompt_printed:
                            self.queue.put("No data received from serial port.")
                            no_data_prompt_printed = True
                    else:
                        no_data_prompt_printed = False
                        ser.write(b"\r\n")

                self.queue.put("Sending print command...")

                ser.write(b"printenv \r\n")
                time.sleep(1)

                response = ser.read_all().decode()

            mac_start_index = response.find("ethaddr=")
            if mac_start_index != -1:
                mac_end_index = response.find("\n", mac_start_index)
                mac_address_line = response[mac_start_index:mac_end_index].strip()
                mac_address = mac_address_line.split("=")[1].replace(":", "").upper()

                self.queue.put(f"Extracted MAC Address: {mac_address}")
                self.process_mac(mac_address, com_port, wb)
            else:
                self.queue.put("MAC address not found in the printenv response.")

        except Exception as e:
            self.queue.put(f"Error: {str(e)}")
        finally:
            self.toggle_buttons()

if __name__ == "__main__":
    gui = APNamerGUI()
    gui.mainloop()