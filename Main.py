import serial
import time
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
import subprocess
import customtkinter as ctk
import tkinter as tk
import threading
import re
import sys

# Define CustomTkinter class with necessary methods and attributes
class CustomTkinter:
    @staticmethod
    def set_appearance_mode(mode):
        CustomTkinter.appearance_mode = mode

    @staticmethod
    def set_default_color_theme(theme):
        CustomTkinter.default_color_theme = theme

    @staticmethod
    def get_color(key):
        if CustomTkinter.appearance_mode == "dark":
            if key == "background":
                return "#121212"
            elif key == "foreground":
                return "#FFFFFF"
            elif key == "accent":
                return "#2196F3"
        else:
            if key == "background":
                return "#FFFFFF"
            elif key == "foreground":
                return "#000000"
            elif key == "accent":
                return "#3f51b5"

    appearance_mode = "default"
    default_color_theme = "default"


# Define APNamerGUI class inheriting from ctk.CTk
class APNamerGUI(ctk.CTk):

    WIDTH = 800
    HEIGHT = 900 

    def __init__(self):
        super().__init__()
        self.title("AP Namer")
        self.initialize_main_window()

    def initialize_main_window(self):
        self.geometry(f"{APNamerGUI.WIDTH}x{APNamerGUI.HEIGHT}")
        CustomTkinter.set_appearance_mode("dark")
        CustomTkinter.set_default_color_theme("red")


        self.tabControl = ctk.CTkTabview(self)
        self.tabControl.pack(expand=1, fill="both")

        self.tab_1 = self.tabControl.add("Info")
        self.tab_2 = self.tabControl.add("Provision")
        self.tab_3 = self.tabControl.add("Manual Port")

        self.com_port = None
        self.file_path = None  # Store the current Excel file path
        
        self.Disabled = False #Toggle function
        self.Override_Used = False # has the COM port been overriden 

        intro_text = """
        Welcome to AP Namer!

        1. Plug in serial cable to computer
        2. Select Excel file in next tab
        3. Plug serial cable into AP
        4. Follow prompts, plug AP into power
        5. Click restart when you want to provision the next AP, it will use the SAME port
        
        *Clicking on the window during provisioning will cause a "Not responding" Windows prompt until the AP is provisioned
        *Use the manual port tab if AruProTool is not finding the COM port you have plugged in

        *Purge tab is to remove prior configurations from APs, if provisioning, that is implemented as part of the naming process.
        """

        manual_override_text = """
        This is for if the program is trying to use the incorrect COM port or one that is already in use.

        Windows users should type COMx, eg. COM1,COM2 - all uppercase, no whitespaces.

        This does not forcefully open a COM port, rather it allows you to select a port manually. Check device manager to see 
        appropriate ports named:
        USB Serial...(COMx) --AP 5xx series 
        or
        Prolific USB-to-Serial Comm Port(COMx) --AP 3xx series

        """
        #Insert text box for 1st tab
        intro_textbox = ctk.CTkTextbox(self.tab_1, wrap=tk.WORD, height=20)
        intro_textbox.insert(tk.END, intro_text)
        intro_textbox.configure(state=tk.DISABLED)
        intro_textbox.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
       # Insert text box for 3rd tab
        manual_override_textbox = ctk.CTkTextbox(self.tab_3, wrap=tk.WORD, height=10)
        manual_override_textbox.insert(tk.END, manual_override_text)
        manual_override_textbox.configure(state=tk.DISABLED)
        manual_override_textbox.pack(fill=tk.BOTH, expand=True, padx=20, pady=(20, 0), side=tk.TOP)  # Pack at the top with padding

        # Calculate half the window width
        half_width = self.tab_3.winfo_reqwidth() // 2

        # Calculate the center of the window
        x_center = half_width // 2

        override_frame = ctk.CTkFrame(self.tab_3)
        override_frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER, y=half_width, x = x_center)  # Place at the center below the textbox

        self.excel_button = ctk.CTkButton(self.tab_2, text="Select Excel File", command=self.select_excel_file)
        self.excel_button.pack(pady=4)
        
        self.restart_button = ctk.CTkButton(self.tab_2, text="Restart", command= self.restart)
        self.restart_button.pack(pady=4)
        
        self.output_textbox = ctk.CTkTextbox(self.tab_2, wrap=tk.WORD, height=20)
        
        self.output_textbox.configure(state=tk.NORMAL)
        self.output_textbox.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        ##Input dialog/button for 3rd tab
        self.override_button = ctk.CTkButton(override_frame, text="Override", command=self.input)
        self.override_button.pack()  # Pack in the center
    
        self.com_port_thread = threading.Thread(target=self.check_com_ports)
        self.com_port_thread.daemon = True
        self.com_port_thread.start()
    
    def toggleButtons(self, tab_1, tab_2, tab_3):
        if self.Disabled:
            self.restart_button.configure(state=tk.ACTIVE)
            self.excel_button.configure(state=tk.ACTIVE)
            self.tabControl.configure(tab_2, state="normal")
            self.tabControl.configure(tab_1, state="normal")
            self.tabControl.configure(tab_3, state="normal")
            self.Disabled = False
        else:
            self.excel_button.configure(state=tk.DISABLED)
            self.restart_button.configure(state=tk.DISABLED)
            self.tabControl.configure(tab_1, state="disabled")
            self.tabControl.configure(tab_2, state="disabled")
            self.tabControl.configure(tab_3, state="disabled")
            self.Disabled = True
            
    def input(self):
            dialog = ctk.CTkInputDialog(text="Type in COM port", title="Override Dialog",
		        fg_color="gray",
		        button_fg_color="blue",
		        button_hover_color="green",
		        button_text_color="black",
		        entry_fg_color="dark gray",
		        entry_border_color="black",
		        entry_text_color="black"
            )
            self.com_port = dialog.get_input()
            self.Override_Used = True
            self.update_output("Overridden, you are now using: " + self.com_port)

    def restart(self):
        # Clear the output textbox
        self.output_textbox.delete(1.0, tk.END)
        # Restart the application with the same Excel file and COM port
        #Verify user has selected Excel file and that a COM port has been decided
        if self.com_port == None:
                self.input()
        if self.file_path == None:
            self.select_excel_file()

        self.start_serial(self.com_port, self.file_path)

    def select_excel_file(self):
        file_path = filedialog.askopenfilename(title="Select Excel file", filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            self.output_textbox.delete(1.0, tk.END)
            self.update_output("Selected Excel file: " + file_path)
            self.file_path = file_path  # Update the stored file path
            self.process_excel_file(file_path)

    def update_output(self, message):
        self.output_textbox.insert(tk.END, message + '\n')
        self.output_textbox.see(tk.END)

    def check_com_ports(self):
        if self.Override_Used:
            return
        first_instance_find_com = True
        first_instance_no_com=True
        while True:
            self.com_port = self.find_com_port()
            if self.com_port and first_instance_find_com:
                self.update_output("COM Port found: " + self.com_port)
                #self.purge_status_label.configure(text="COM Port: " + self.com_port)
                first_instance_find_com = False
            if self.com_port:
                break
            if first_instance_no_com == True:
                self.update_output("No COM Port found")
                #self.purge_status_label.configure(text="COM Port: Not Detected")
                first_instance_no_com = False
            else: 
                pass

    def process_excel_file(self, file_path):
        if self.com_port:
            self.start_serial(self.com_port, file_path)


    #Use powershell script to find COM port and strip result to use 'COMx'
    def find_com_port(self):
        if sys.platform.startswith('win32'):
            powershell_cmd = [
            'powershell.exe',
            '-Command',
            '''
            $5xxcomPort = Get-CimInstance Win32_PnPEntity | Where-Object { $_.Caption -like '*USB Serial Port*' }
            $3xxcomPort = Get-CimInstance Win32_PnPEntity | Where-Object { $_.Caption -like '*Prolific USB-to-Serial Comm Port*' }

            if ($5xxcomPort) {
                $5xxcomPortName = $5xxcomPort.Name
                Write-Output "USB Serial Port found on COM port: $5xxcomPortName"
            } elseif ($3xxcomPort) {
                $3xxcomPortName = $3xxcomPort.Name
                Write-Output "USB Serial Port found on COM port: $3xxcomPortName"
            } else {
                Write-Output "USB Serial Port not found"
            }
            '''
            ]
            #Strip result logic
            startupinfo = subprocess.STARTUPINFO()
            startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
            try:
                result = subprocess.check_output(powershell_cmd, text=True, startupinfo=startupinfo)
                lines = result.strip().split('\n')
                for line in lines:
                    if "USB Serial Port" in line:
                        com_port_index = line.find("(COM")
                        if com_port_index != -1:
                            com_port = line[com_port_index:].strip("() ")
                            return com_port
                else:
                    return None
            except subprocess.CalledProcessError as e:
                print(f"Error: {e}")
                return None
        else:
            time.sleep(1)
            pass

    #Look for columns MAC, AP Name, and AP Group
    #Look for MAC row, give AP {name} {group} in MAC row, save 
    #Give the user the results
    def process_mac(self, mac_to_match, com_port, wb):
        sheet = wb.active
        headers = {}
        for cell in sheet[1]:
            if cell.value:
                headers[cell.value] = cell.column

        # Edit column names here
        mac_column_index = headers.get("MAC")
        name_column_index = headers.get("AP Name")
        group_column_index = headers.get("AP Group")

        if mac_column_index is None or name_column_index is None or group_column_index is None:
            self.update_output("Error: Required column not found in Excel file.")
            return

        for row in sheet.iter_rows(min_row=2, values_only=True):
            print("Row:", row)
            if row[mac_column_index - 1] == mac_to_match and row[mac_column_index - 1] is not None:
                name = row[name_column_index - 1]
                group = row[group_column_index - 1]

                #Configuration of serial is based off: https://community.arubanetworks.com/discussion/console-port-on-ap-515-no-response 
                with serial.Serial(com_port, baudrate=9600, timeout=1, parity=serial.PARITY_NONE, bytesize=8, stopbits=1, xonxoff=False, rtscts=False, dsrdtr=False) as ser:
                    ser.write(f"purgeenv\r\n".encode())
                    time.sleep(1)

                    ser.write(f"saveenv\r\n".encode())
                    time.sleep(1)

                    ser.write(f"set name {name}\r\n".encode())
                    time.sleep(1)

                    ser.write(f"set group {group}\r\n".encode())
                    time.sleep(1)

                    ser.write(b"saveenv \r\n")
                    self.update_output("Sent commands...")
                    time.sleep(2)

                    ser.write(b"printenv \r\n")
                    time.sleep(1)

                    # Give provided name/group/mac : source : AP
                    response = ser.read_all().decode()

                    match = re.search(r'\bname=([^\n]+)', response)
                    if match:
                        actualName = match.group(1)

                    match = re.search(r'\bgroup=([^\n]+)', response)
                    if match:
                        actualGroup = match.group(1)

                    match = re.search(r'\bethaddr=([^\n]+)', response)
                    if match:
                        actualMac = match.group(1)

                    # Pass name, group, MAC and activate buttons
                    self.toggleButtons(self.tab_1, self.tab_2, self.tab_3)
                    self.update_output("Extracted name: " + actualName + "\n" + "Extracted group: " + actualGroup + "\n" + "Extracted MAC: " + actualMac)
                    self.update_output("\n\n AP Successfully Provisioned.")
                break
        else:
            self.update_output("MAC address not found in the Excel file.")
            self.toggleButtons(self.tab_1, self.tab_2, self.tab_3)

    def start_serial(self, com_port, file_path):
        
        try:
            wb = load_workbook(file_path, read_only=True)
            self.file_path = file_path  # Store the file path
            self.update_idletasks()  # Force GUI update
            self.toggleButtons(self.tab_1, self.tab_2, self.tab_3)
            

            with serial.Serial(com_port, baudrate=9600, timeout=1) as ser:
                no_data_prompt_printed = False

                while True:
                    data = ser.readline()

                    if b"Hit <Enter>" in data:
                        self.update_output(data.decode())
                        ser.write(b"\r\n")
                        break
                    elif not data:
                        if not no_data_prompt_printed:
                            self.update_output("No data received from serial port.")
                            self.update_idletasks()  # Force GUI update
                            no_data_prompt_printed = True
                    else:
                        no_data_prompt_printed = False
                        ser.write(b"\r\n")

                self.update_output("Sending print command...")
                self.update_idletasks()  # Force GUI update

                ser.write(b"printenv \r\n")
                time.sleep(1)

                response = ser.read_all().decode()

                ser.close()  #This will not work without this line. It will 're-open' the already being used COM port.

                mac_start_index = response.find("ethaddr=")
                if mac_start_index != -1:
                    mac_end_index = response.find("\n", mac_start_index)
                    mac_address_line = response[mac_start_index:mac_end_index].strip()
                    mac_address = mac_address_line.split("=")[1].replace(":", "").upper()

                    self.update_output("Extracted MAC Address:" + mac_address)
                    self.process_mac(mac_address, com_port, wb)
                    
                else:
                    self.update_output("MAC address not found in the printenv response.")
                    self.toggleButtons(self.tab_1, self.tab_2, self.tab_3)

        except Exception as e:
            self.update_output(f"Error: {str(e)}")
            self.toggleButtons(self.tab_1, self.tab_2, self.tab_3)


if __name__ == "__main__":
        gui = APNamerGUI()
        gui.mainloop()
